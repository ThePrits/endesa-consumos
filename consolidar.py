"""
Consolidador de facturas de consumo eléctrico.

Uso:
    python consolidar.py

Lee todos los CSV de facturas en la carpeta 'input/' y genera
'output/consolidado_consumo.csv' con todos los registros horarios
ordenados cronológicamente.

Detecta y elimina archivos duplicados (mismo contenido de consumo)
antes de consolidar. Los duplicados se mueven a 'input/_duplicados/'.
"""

from __future__ import annotations

import csv
import hashlib
import os
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    XLSX_DISPONIBLE = True
except ImportError:
    XLSX_DISPONIBLE = False

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------

INPUT_DIR = Path(__file__).resolve().parent / "input"
OUTPUT_DIR = Path(__file__).resolve().parent / "output"
ARCHIVO_SALIDA = "consolidado_consumo.csv"
ARCHIVO_SALIDA_XLSX = "consolidado_consumo.xlsx"
CARPETA_DUPLICADOS = "_duplicados"

# Separador y decimal del archivo de salida
SEPARADOR_SALIDA = ";"
DECIMAL_SALIDA = ","

ENCODING = "utf-8-sig"


# ---------------------------------------------------------------------------
# Estructuras de datos
# ---------------------------------------------------------------------------

@dataclass
class Factura:
    ruta: Path
    nombre: str
    fecha_inicio: datetime
    fecha_fin: datetime
    hash_consumo: str
    num_registros: int
    es_duplicado: bool = False
    duplicado_de: str | None = None
    registros: list[dict] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Parseo de fechas y normalización
# ---------------------------------------------------------------------------

def parse_fecha_ddmmyyyy(texto: str) -> datetime:
    return datetime.strptime(texto.strip(), "%d/%m/%Y")


def parse_fecha_iso(texto: str) -> datetime:
    return datetime.strptime(texto.strip(), "%Y-%m-%d")


def formato_fecha_salida(dt: datetime) -> str:
    return dt.strftime("%d/%m/%Y")


def formato_numero(valor: float) -> str:
    return f"{valor:.2f}".replace(".", DECIMAL_SALIDA)


def normalizar_hora(hora: str) -> str:
    return hora.strip().replace("-", " - ")


def normalizar_consumo(valor: str) -> float:
    texto = valor.strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    return round(float(texto), 2)


# ---------------------------------------------------------------------------
# Lectura de CSV
# ---------------------------------------------------------------------------

def leer_filas(ruta: Path) -> list[list[str]]:
    """Lee todas las filas de un CSV sin procesar."""
    with ruta.open(newline="", encoding=ENCODING) as f:
        return list(csv.reader(f))


def calcular_hash(filas: list[list[str]]) -> tuple[str, int]:
    """
    Genera una huella SHA-256 basada en Fecha + Hora + Consumo.
    Dos archivos con los mismos datos de consumo producen el mismo hash,
    independientemente de su nombre o metadatos.
    """
    registros: list[str] = []

    for fila in filas[6:]:
        if not fila or len(fila) < 3:
            continue
        if fila[0].strip().lower() == "fecha":
            continue
        try:
            fecha = parse_fecha_iso(fila[0]).strftime("%Y-%m-%d")
            hora = normalizar_hora(fila[1])
            consumo = f"{normalizar_consumo(fila[2]):.2f}"
        except (ValueError, IndexError):
            continue
        registros.append(f"{fecha}|{hora}|{consumo}")

    registros.sort()
    payload = "\n".join(registros).encode("utf-8")
    return hashlib.sha256(payload).hexdigest(), len(registros)


def leer_factura(ruta: Path) -> Factura:
    """
    Lee los metadatos y calcula el hash de una factura CSV.

    Estructura esperada del CSV:
        Fila 1: CUPS, <valor>
        Fila 2: Fecha inicio:, <DD/MM/YYYY>
        Fila 3: Fecha fin:,    <DD/MM/YYYY>
        Fila 4-5: otros metadatos
        Fila 6: línea vacía
        Fila 7: cabecera (Fecha, Hora, Consumo (Wh), ...)
        Fila 8+: registros horarios (YYYY-MM-DD, HH:MM-HH:MM, Wh, ...)
        Última fila: fila de total (,Total (Wh):, <valor>)
    """
    filas = leer_filas(ruta)

    if len(filas) < 3:
        raise ValueError("El archivo tiene menos de 3 filas.")

    try:
        fecha_inicio = parse_fecha_ddmmyyyy(filas[1][1])
        fecha_fin = parse_fecha_ddmmyyyy(filas[2][1])
    except (IndexError, ValueError) as exc:
        raise ValueError(f"No se pudieron leer las fechas de metadatos: {exc}") from exc

    hash_consumo, num_registros = calcular_hash(filas)

    if num_registros == 0:
        raise ValueError("No se encontraron registros horarios válidos.")

    return Factura(
        ruta=ruta,
        nombre=ruta.name,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        hash_consumo=hash_consumo,
        num_registros=num_registros,
    )


def leer_registros(factura: Factura) -> list[dict]:
    """Lee los registros horarios de una factura ya cargada."""
    filas = leer_filas(factura.ruta)
    registros = []

    for fila in filas[6:]:
        if not fila or len(fila) < 3:
            continue
        if fila[0].strip().lower() == "fecha":
            continue
        try:
            fecha_dt = parse_fecha_iso(fila[0])
            hora = normalizar_hora(fila[1])
            consumo_wh = normalizar_consumo(fila[2])
        except (ValueError, IndexError):
            continue

        registros.append({
            "fecha_dt": fecha_dt,
            "fecha": formato_fecha_salida(fecha_dt),
            "hora": hora,
            "consumo_wh": consumo_wh,
            "archivo_origen": factura.nombre,
        })

    return registros


# ---------------------------------------------------------------------------
# Detección y gestión de duplicados
# ---------------------------------------------------------------------------

def detectar_duplicados(facturas: list[Factura]) -> tuple[list[Factura], list[Factura]]:
    """
    Agrupa las facturas por hash de consumo.

    De cada grupo se conserva la factura con la fecha de inicio más
    temprana; en caso de empate, la de nombre alfabéticamente menor.
    Devuelve (únicas, duplicadas).
    """
    por_hash: dict[str, list[Factura]] = {}
    for f in facturas:
        por_hash.setdefault(f.hash_consumo, []).append(f)

    unicas: list[Factura] = []
    duplicadas: list[Factura] = []

    for grupo in por_hash.values():
        grupo.sort(key=lambda f: (f.fecha_inicio, f.fecha_fin, f.nombre.lower()))
        principal = grupo[0]
        unicas.append(principal)
        for dup in grupo[1:]:
            dup.es_duplicado = True
            dup.duplicado_de = principal.nombre
            duplicadas.append(dup)

    return unicas, duplicadas


def mover_duplicados(duplicadas: list[Factura]) -> None:
    """Mueve los archivos duplicados a input/_duplicados/."""
    if not duplicadas:
        return

    destino_dir = INPUT_DIR / CARPETA_DUPLICADOS
    destino_dir.mkdir(exist_ok=True)

    for dup in duplicadas:
        destino = destino_dir / dup.nombre
        # Evitar colisión si ya existe un archivo con ese nombre en destino
        contador = 2
        while destino.exists():
            destino = destino_dir / f"{dup.ruta.stem}_{contador}{dup.ruta.suffix}"
            contador += 1
        shutil.move(str(dup.ruta), str(destino))
        print(f"  Duplicado movido: {dup.nombre} → {destino.relative_to(INPUT_DIR.parent)}")


# ---------------------------------------------------------------------------
# Consolidación y escritura
# ---------------------------------------------------------------------------

def consolidar(facturas: list[Factura]) -> list[dict]:
    """
    Une todos los registros horarios de las facturas en una sola lista,
    elimina duplicados exactos (misma fecha + hora + archivo_origen)
    y ordena de más reciente a más antiguo.
    """
    todos: list[dict] = []
    claves_vistas: set[tuple] = set()
    omitidos = 0

    for factura in facturas:
        for reg in leer_registros(factura):
            clave = (reg["fecha"], reg["hora"], reg["archivo_origen"])
            if clave in claves_vistas:
                omitidos += 1
                continue
            claves_vistas.add(clave)
            todos.append(reg)

    todos.sort(key=lambda r: (r["fecha_dt"], r["hora"]), reverse=True)

    if omitidos:
        print(f"  Registros duplicados omitidos dentro del consolidado: {omitidos}")

    return todos


def guardar_consolidado(registros: list[dict]) -> Path:
    """Escribe el CSV consolidado en output/."""
    OUTPUT_DIR.mkdir(exist_ok=True)
    ruta_salida = OUTPUT_DIR / ARCHIVO_SALIDA

    with ruta_salida.open("w", newline="", encoding=ENCODING) as f:
        writer = csv.writer(f, delimiter=SEPARADOR_SALIDA)
        writer.writerow(["Fecha", "Hora", "Consumo_Wh", "Archivo_origen"])
        for reg in registros:
            writer.writerow([
                reg["fecha"],
                reg["hora"],
                formato_numero(reg["consumo_wh"]),
                reg["archivo_origen"],
            ])

    return ruta_salida


def guardar_xlsx(registros: list[dict]) -> Path:
    """
    Escribe el consolidado en formato Excel (.xlsx) en output/.

    Formato aplicado:
    - Cabecera en negrita con fondo azul oscuro y texto blanco.
    - Columna Consumo_Wh como número con dos decimales.
    - Columna Fecha como fecha real de Excel (DD/MM/YYYY).
    - Ancho de columnas ajustado al contenido.
    - Fila de total de consumo al final.
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    ruta_salida = OUTPUT_DIR / ARCHIVO_SALIDA_XLSX

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    # -- Estilos --
    fondo_cabecera = PatternFill("solid", fgColor="1F4E79")
    fuente_cabecera = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    fuente_datos = Font(name="Arial", size=10)
    fuente_total = Font(name="Arial", bold=True, size=10)
    alineacion_centro = Alignment(horizontal="center", vertical="center")
    alineacion_der = Alignment(horizontal="right", vertical="center")

    fmt_fecha = "DD/MM/YYYY"
    fmt_consumo = '#,##0.00'

    # -- Cabecera --
    cabeceras = ["Fecha", "Hora", "Consumo_Wh", "Archivo_origen"]
    for col, titulo in enumerate(cabeceras, start=1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.font = fuente_cabecera
        celda.fill = fondo_cabecera
        celda.alignment = alineacion_centro

    # -- Datos --
    for fila_idx, reg in enumerate(registros, start=2):
        # Fecha como objeto datetime para que Excel la reconozca como fecha
        ws.cell(row=fila_idx, column=1, value=reg["fecha_dt"]).number_format = fmt_fecha
        ws.cell(row=fila_idx, column=1).font = fuente_datos
        ws.cell(row=fila_idx, column=1).alignment = alineacion_centro

        ws.cell(row=fila_idx, column=2, value=reg["hora"]).font = fuente_datos
        ws.cell(row=fila_idx, column=2).alignment = alineacion_centro

        celda_consumo = ws.cell(row=fila_idx, column=3, value=reg["consumo_wh"])
        celda_consumo.number_format = fmt_consumo
        celda_consumo.font = fuente_datos
        celda_consumo.alignment = alineacion_der

        ws.cell(row=fila_idx, column=4, value=reg["archivo_origen"]).font = fuente_datos

    # -- Fila de total --
    fila_total = len(registros) + 2
    ws.cell(row=fila_total, column=2, value="TOTAL").font = fuente_total
    ws.cell(row=fila_total, column=2).alignment = alineacion_centro

    primera_dato = 2
    ultima_dato = len(registros) + 1
    celda_total = ws.cell(
        row=fila_total,
        column=3,
        value=f"=SUM(C{primera_dato}:C{ultima_dato})",
    )
    celda_total.number_format = fmt_consumo
    celda_total.font = fuente_total
    celda_total.alignment = alineacion_der

    # -- Ancho de columnas --
    anchos = {"A": 13, "B": 16, "C": 14, "D": 45}
    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = ancho

    # -- Inmovilizar cabecera --
    ws.freeze_panes = "A2"

    wb.save(ruta_salida)
    return ruta_salida


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("=" * 60)
    print("Consolidador de consumos eléctricos")
    print("=" * 60)

    # Buscar CSVs en input/, ignorando la subcarpeta de duplicados
    rutas_csv = sorted(
        [
            p for p in INPUT_DIR.glob("*.csv")
            if p.parent == INPUT_DIR  # excluye subdirectorios
        ],
        key=lambda p: p.name.lower(),
    )

    if not rutas_csv:
        print(f"\nNo se encontraron archivos CSV en '{INPUT_DIR}'.")
        print("Coloca las facturas CSV descargadas de tu distribuidora en esa carpeta.")
        return

    print(f"\nArchivos encontrados: {len(rutas_csv)}")

    # Leer metadatos de cada factura
    facturas: list[Factura] = []
    for ruta in rutas_csv:
        try:
            facturas.append(leer_factura(ruta))
            print(f"  OK  {ruta.name}")
        except Exception as exc:
            print(f"  OMITIDO  {ruta.name} — {exc}")

    if not facturas:
        print("\nNo se pudo procesar ningún archivo. Revisa el formato de los CSV.")
        return

    # Detectar duplicados
    unicas, duplicadas = detectar_duplicados(facturas)

    if duplicadas:
        print(f"\nDuplicados detectados: {len(duplicadas)}")
        mover_duplicados(duplicadas)
    else:
        print("\nNo se detectaron duplicados.")

    # Ordenar por fecha de inicio
    unicas.sort(key=lambda f: (f.fecha_inicio, f.fecha_fin, f.nombre.lower()))

    # Consolidar y guardar
    print(f"\nConsolidando {len(unicas)} factura(s)...")
    registros = consolidar(unicas)
    ruta_csv = guardar_consolidado(registros)

    ruta_xlsx = None
    if XLSX_DISPONIBLE:
        ruta_xlsx = guardar_xlsx(registros)
    else:
        print("  Aviso: openpyxl no está instalado. Se omite la exportación a XLSX.")
        print("         Instálalo con: pip install openpyxl")

    # Resumen final
    print("\n" + "=" * 60)
    print("Proceso completado")
    print(f"  Facturas procesadas : {len(unicas)}")
    print(f"  Duplicados movidos  : {len(duplicadas)}")
    print(f"  Registros totales   : {len(registros)}")
    print(f"  CSV generado        : {ruta_csv}")
    if ruta_xlsx:
        print(f"  XLSX generado       : {ruta_xlsx}")
    print("=" * 60)


if __name__ == "__main__":
    main()